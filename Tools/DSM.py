###############################################
#                                             #
# Design Structure Matrix Module              #
#                                             #
# simple module for creating and managing DSM #
# Contrib: uChouinard                         #
# V0 2/12/2016                                #
#                                             #
###############################################

#%matplotlib inline
import numpy as np
import matplotlib.pyplot as plt
import networkx as nx
#import warnings
import pandas as pds
import genFN as fzn
from openpyxl import load_workbook


        
class DSM():
    
    def __init__(self, name='none', dsmType='simple', directed='yes'):
        
        
        self.name=name
        self.complist=[]
        
        if directed is 'yes':
            if dsmType is 'simple':
                self.dsm=nx.DiGraph()
            elif dsmType is 'interactions':
                self.dsm={'s':nx.DiGraph(),'i': nx.DiGraph(),
                      'e':nx.DiGraph(), 'm':nx.DiGraph()}
            elif dsmType is 'dependencies':
                self.dsm={'v':nx.DiGraph(),'h': nx.DiGraph(),
                      'e':nx.DiGraph()}
            elif dsmType is 'functional':
                self.dsm={'i':nx.DiGraph(),'e': nx.DiGraph()}
            self.twoDdsm=nx.DiGraph()
            self.real2D=nx.DiGraph()
            self.imag2D=nx.DiGraph()
        else:
            if dsmType is 'simple':
                self.dsm=nx.Graph()
            elif dsmType is 'interactions':
                self.dsm={'s':nx.Graph(),'i': nx.Graph(),
                      'e':nx.Graph(), 'm':nx.Graph()}
            elif dsmType is 'dependencies':
                self.dsm={'v':nx.Graph(),'h': nx.Graph(),
                      'e':nx.Graph()}
            elif dsmType is 'functional':
                self.dsm={'i':nx.Graph(),'e': nx.Graph()}
            self.twoDdsm=nx.Graph()
            self.real2D=nx.Graph()
            self.imag2D=nx.Graph()
        self.dsmType=dsmType
       
    def getDSM(self):
        return self.dsm
    
    def addComponent(self, cName):
        
        for name in cName:
            self.complist.append(name)
        
            if self.dsmType=='simple':
                self.dsm.add_node(name)
            elif self.dsmType=='interactions':
                for keys in self.dsm:
                    self.dsm[keys].add_node(name)
            elif self.dsmType=='dependencies':
                for keys in self.dsm:
                    self.dsm[keys].add_node(name)
            elif self.dsmType=='functional':
                for keys in self.dsm:
                    self.dsm[keys].add_node(name) 
        
    def addRelation(self, cNameFrom, cNameTo, val):
        
        if len(cNameFrom) == len(cNameTo):
            for i in range(len(cNameFrom)):
                if not(cNameFrom[i] in self.complist):
                    self.addComponent([cNameFrom[i]])
            
                if not(cNameTo[i] in self.complist):
                    self.addComponent([cNameTo[i]])
                
            
                if self.dsmType=='simple':
            #if type(val) is float:
                    self.dsm.add_edge(cNameFrom[i], cNameTo[i], weight=val[i])
            #else :
            #    warnings.warn('non valid input value')
        
                if self.dsmType== 'interactions':
            #if type(val) is interactions:
                    for keys in val[i]:
                        self.dsm[keys].add_edge(cNameFrom[i], cNameTo[i], weight=val[i][keys])
                if self.dsmType== 'dependencies':
            #if type(val) is dependencies:
                    for keys in val[i]:
                        self.dsm[keys].add_edge(cNameFrom[i], cNameTo[i], weight=val[i][keys])
                        
                if self.dsmType== 'functional':
            #if type(val) is dependencies:
                    for keys in val[i]:
                        self.dsm[keys].add_edge(cNameFrom[i], cNameTo[i], weight=val[i][keys])
    def getRelation(self, a,b):
        res={}
        for key in self.dsm:
            #print (key)
            res[key]=self.dsm[key].get_edge_data(a,b)
            #print(self.dsm[key].get_edge_data(a,b))
        return res
    def display(self):
        print (self.complist)
        if self.dsmType == 'simple':
            
            print (nx.to_numpy_matrix(self.dsm, nodelist=self.complist))
        
        elif self.dsmType == 'interactions':
            
            for keys in ['s','e','i','m']:
                print (keys)
                print (nx.to_numpy_matrix(self.dsm[keys], nodelist=self.complist)  )
                
        elif self.dsmType == 'dependencies':
            
            for keys in ['v','h','e']:
                print (keys)
                print (nx.to_numpy_matrix(self.dsm[keys], nodelist=self.complist)  )
                
        elif self.dsmType == 'functional':
            
            for keys in ['i','e']:
                print (keys)
                print (nx.to_numpy_matrix(self.dsm[keys], nodelist=self.complist)  ) 
                
    def to_Pandas_Data_Frame(self, _type='regular'):
        
        if self.dsmType == 'simple':
            df=pds.DataFrame(data=nx.to_numpy_matrix(self.dsm, nodelist=self.complist), columns=self.complist, index=self.complist)
              
        
        if self.dsmType == 'interactions':
            
            x,y = len(self.complist),len(self.complist)
            A = [ ['']*x for i in range(y) ]
            k=0
            
            if _type is 'regular':
                
                for keys in ['s','e','i','m']:
                    weights=nx.get_edge_attributes(self.dsm[keys],'weight')
                    #myGmat=nx.to_numpy_matrix(self.dsm[keys], nodelist=self.complist) 
                    for pair in weights.keys():
                        #print(pair)
                        i=self.complist.index(pair[0])
                        j=self.complist.index(pair[1])
                        A[i][j]=A[i][j]+str(keys)+':'+ str(weights[pair])+'/  '
                
                
            elif _type is '2D':
                weights=nx.get_edge_attributes(self.twoDdsm,'weight')
                for pair in weights.keys():
                        i=self.complist.index(pair[0])
                        j=self.complist.index(pair[1])
                        A[i][j]=str(weights[pair])
                 
            df=pds.DataFrame(data=A, columns=self.complist, index=self.complist)
            
        if self.dsmType == 'dependencies':
            
            x,y = len(self.complist),len(self.complist)
            A = [ ['']*x for i in range(y) ]
            k=0
            for keys in ['v','h', 'e']:
                weights=nx.get_edge_attributes(self.dsm[keys],'weight')
                for pair in weights.keys():
                        i=self.complist.index(pair[0])
                        j=self.complist.index(pair[1])
                        A[i][j]=A[i][j]+keys+':'+ str(weights[pair])+'/  '
                            
            df=pds.DataFrame(data=A, columns=self.complist, index=self.complist)
        return df
    
    def toExcel(self, filename):
            
        if self.dsmType == 'simple':
            df=pds.DataFrame(data=nx.to_numpy_matrix(self.dsm, nodelist=self.complist), columns=self.complist, index=self.complist)
            
            book = load_workbook(filename)
            writer = pds.ExcelWriter(filename, engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            
            df.to_excel(writer,'simpleDSM')
            writer.save()
        
        if self.dsmType == 'interactions':
            
            sheetList={'s':'Spatial_DSM', 'e':'Energy_DSM', 'i':'Information_DSM', 'm':'Material_DSM'}
            
            x,y = len(self.complist),len(self.complist)
            A = [ ['']*x for i in range(y) ]
            k=0
            
            book = load_workbook(filename)
            writer = pds.ExcelWriter(filename, engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            for keys in ['s','e','i','m']:
                myGmat=nx.to_numpy_matrix(self.dsm[keys], nodelist=self.complist)  
                 
                df=pds.DataFrame(data=myGmat, columns=self.complist, index=self.complist)
                
                df.to_excel(writer,sheetList[keys])
            writer.save()
            
        if self.dsmType == 'dependencies':
            sheetList={'v':'Vibration_DSM', 'h':'Heat_DSM', 'e':'EMF_DSM'}
            
            x,y = len(self.complist),len(self.complist)
            A = [ ['']*x for i in range(y) ]
            k=0
            
            book = load_workbook(filename)
            writer = pds.ExcelWriter(filename, engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            for keys in ['v','h', 'e']:
                myGmat=nx.to_numpy_matrix(self.dsm[keys], nodelist=self.complist)  
                    
                df=pds.DataFrame(data=myGmat, columns=self.complist, index=self.complist)
                df.to_excel(writer,sheetList[keys])
            writer.save()
            
    def to2Dinteractions(self, weights):
        
        A_complex=(np.zeros((len(self.complist), len(self.complist)))).astype(complex)
        A_real=np.zeros(A_complex.shape)
        A_imag=np.zeros(A_complex.shape)
        for keys in ['s','e','i','m']:
            tmpMat=weights[keys]*nx.to_numpy_matrix(self.dsm[keys], nodelist=self.complist) 
            tmpComplex=tmpMat.astype(complex)
            tmpReal=np.copy(tmpMat)
            tmpReal[(tmpReal<0)]=0
            tmpImag=np.copy(tmpMat)
            tmpImag[tmpImag>0]=0
            tmpImag=np.abs(tmpImag)
            tmpComplex=tmpReal+tmpImag*1.j
            #print(abs(tmpMat[tmpMat<0])*1.j)
            #flatTmp=tmpMat.flatten()
            #for row in range(len(self.complist)):
            #    for col in range(len(self.complist)):
            #        if tmpMat[row,col]<0:
            #            tmpComplex[row,col]=0+abs(tmpMat[row,col])*1.j
                    
            
            A_complex=A_complex+tmpComplex
            A_real=A_real+tmpReal
            A_imag=A_imag+tmpImag
            
        #print(A_complex)
        #print(A_real)
        #print(A_imag)
        if nx.is_directed(self.twoDdsm):
            self.twoDdsm=nx.DiGraph(A_complex)
            self.real2D=nx.DiGraph(A_real)
            self.imag2D=nx.DiGraph(A_imag)
        else:
            self.twoDdsm=nx.Graph(A_complex)
            self.real2D=nx.Graph(A_real)
            self.imag2D=nx.Graph(A_imag)
            
        mapping=dict(zip(self.twoDdsm.nodes(),self.complist ))
        self.twoDdsm=nx.relabel_nodes(self.twoDdsm, mapping)
        self.real2D=nx.relabel_nodes(self.real2D, mapping)
        self.imag2D=nx.relabel_nodes(self.imag2D, mapping)
