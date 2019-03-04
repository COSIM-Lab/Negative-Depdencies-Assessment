###############################################
#                                             #
# Interfacing with Excel Module to build DSM  #
#                                             #
# Contrib: uChouinard                         #
# V0 03/03/2019                               #
#                                             #
###############################################

import DSM as dsmx
import SystemDependencies as dpx
import DependencyIndex as dpi
import pandas as pds
from openpyxl import load_workbook

#purely Based on Negative Dependencies DSM
class NDI_Interfacer:
    
    
    def __init__(self, input_filename, output_filename=''):
        
        
        self.input_filename=input_filename
        self.output_filename=''
        if output_filename is '':
            self.output_filename=input_filename
        
        self.sys=sys=dpx.System('')
        
        

    def dsmBuilder(self):
    
        #Fetching Relevant Affecting/Affected Info
        df=pds.read_excel(self.input_filename, 'Input_Level')
        compList=[]
    
        mat_nan=df.isnull()
    
        for _component in list(df['Components'].values):
            tmp_comp=dpx.Component(_component)
            cmp_index=list(df['Components']).index(_component)
        
            if not mat_nan['AD_Heat'][ cmp_index]:
                tmp_comp.addAffected(dpx.AdverseEffect('Heat',df['AD_Heat'][ cmp_index]))
            if not mat_nan['AD_Vibration'][ cmp_index]:    
                tmp_comp.addAffected(dpx.AdverseEffect('Vibration', df['AD_Vibration'][ cmp_index]))
            if not mat_nan['AD_EMF'][ cmp_index]:     
                tmp_comp.addAffected(dpx.AdverseEffect('EMF', df['AD_EMF'][ cmp_index]))
            if not mat_nan['AR_Heat'][ cmp_index]:
                tmp_comp.addAffecting(dpx.AdverseEffect('Heat', df['AR_Heat'][ cmp_index]))
            if not mat_nan['AR_Vibration'][ cmp_index]:
                tmp_comp.addAffecting(dpx.AdverseEffect('Vibration', df['AR_Vibration'][ cmp_index]))
            if not mat_nan['AR_EMF'][ cmp_index]:
                tmp_comp.addAffecting(dpx.AdverseEffect('EMF', df['AR_EMF'][ cmp_index]))
        
            compList.append(tmp_comp)
    
        #Fetching Layout Information
        dfc=pds.read_excel(self.input_filename, 'Input_Closeness')
        mat_nan_c=dfc.isnull()
        closeness={}
        other_components=list(dfc.keys())
        other_components.remove('Components')
        for _component in list(dfc['Components'].values):
        
            cmp_index=list(df['Components']).index(_component)
            for _other_component in other_components:
            
                if not mat_nan_c[_other_component][cmp_index]:
                
                    closeness[('/'.join(sorted([_component,_other_component])))]=dfc[_other_component][cmp_index]
        #print(closeness)
        sys=dpx.System('Drone')
        sys.addComponents(compList)
        sys.addCloseness(closeness)
    
        dfa=pds.read_excel(self.input_filename, 'Input_Attenuation')


        if (not dfa.isnull().any().any()) and (list(dfa.index)):
            att_level=dfa.to_dict('records')
            sys.attenuation=att_level[0]
        
        sys.build('float')
    
        
        sys.adverseDSM.toExcel(self.output_filename)
        
        self.sys=sys
   

	#analysis of the system using the NDI
    def sysAnalysis(self, _indexType='randic',  _normType='linear'):
        dfm=pds.read_excel(self.input_filename, 'Input_Effect_Fuzzy_Measures')
        if (not dfm.isnull().any().any()) and (list(dfm.index)):
            fuzzy_Measures=dfm.to_dict('records')[0]
        else:
            fuzzy_Measures={'ehv':1.0, 'ev':0.75, 'e':0.70, 'v':0.30, 'h':0.50, 'hv':0.55, 'eh':0.85}
            
        NDI=dpi.calcNDI(self.sys.adverseDSM, fuzzy_Measures, len(self.sys.components), _indexType, _normType)
        
        book = load_workbook(self.output_filename)
        writer = pds.ExcelWriter(self.output_filename, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        ndiDict={'NDI':[NDI]}
        #print(ndiDict)
        dfNDI=pds.DataFrame.from_dict(ndiDict)
        dfNDI.to_excel(writer,'NDI_Analysis')
        writer.save()
        
#Based on spatial/information,material/energy DSM
class Interactions_Interfacer:
    
    def __init__(self, input_filename, output_filename='', _directed='yes'):
                
        self.input_filename=input_filename
        self.output_filename=''
        if output_filename is '':
            self.output_filename=input_filename
        
        self.dsm=dsmx.DSM(name='none', dsmType='interactions', directed=_directed)
        
        
    def dsmBuilder(self):
        
        
        dfs=pds.read_excel(self.input_filename, 'Spatial')
        dfe=pds.read_excel(self.input_filename, 'Energy')
        dfi=pds.read_excel(self.input_filename, 'Information')
        dfm=pds.read_excel(self.input_filename, 'Material')
        
        dfs=dfs.fillna(0)
        dfe=dfe.fillna(0)
        dfi=dfi.fillna(0)
        dfm=dfm.fillna(0)
        
        other_components=list(dfs.keys())
        other_components.remove('Components')
        
        for _component in list(dfs['Components'].values):
            
            cmp_index=list(dfs['Components']).index(_component)
            for _other_component in other_components:
                tmp_rel={'s':dfs[_other_component][cmp_index],
                         'e':dfe[_other_component][cmp_index],
                         'i':dfi[_other_component][cmp_index],
                         'm':dfm[_other_component][cmp_index]}
                #print(tmp_rel)
                self.dsm.addRelation([_component], [_other_component], [tmp_rel])
                
    def sysAnalysis(self, _indexType='2D',_calculation_method='randic' ,_weights={'s':1, 'e':1, 'i':1, 'm':1}):
        
        res_aggregation=0
        res_index=0
        if _indexType is '2D':
            self.dsm.to2Dinteractions(_weights)
            
            if _calculation_method is 'energy':
                res_real=dpi.energyIndex(self.dsm.real2D)
                res_imag=dpi.energyIndex(self.dsm.imag2D)
            if _calculation_method is 'randic':
                res_real=dpi.randicIndex(self.dsm.real2D)
                res_imag=dpi.randicIndex(self.dsm.imag2D)
            if _calculation_method is 'wiener':
                res_real=dpi.wienerIndex(self.dsm.real2D)
                res_imag=dpi.wienerIndex(self.dsm.imag2D)
            if _calculation_method is 'sum_connectivity':
                res_real=dpi.scIndex(self.dsm.real2D)
                res_imag=dpi.scIndex(self.dsm.imag2D)
            #print(res_aggregation)
            #res_index=dpi.calc2DIndex(res_aggregation.real, res_aggregation.imag)
            res_index=dpi.calc2DIndex(res_real, res_imag)
            #print(res_index)
        
            book = load_workbook(self.output_filename)
            writer = pds.ExcelWriter(self.output_filename, engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
            ndiDict={'2D_Index':[res_index]}
            #print(ndiDict)
            dfNDI=pds.DataFrame.from_dict(ndiDict)
            dfNDI.to_excel(writer,'SIEI')
            
            twoDdsm=self.dsm.to_Pandas_Data_Frame('2D')
            twoDdsm.to_excel(writer,'2D_DSM')
            writer.save()    